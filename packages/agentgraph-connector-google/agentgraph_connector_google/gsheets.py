"""Google Sheets connector."""

# pyright: reportUnknownMemberType=false, reportUnknownVariableType=false
# pyright: reportUnknownArgumentType=false

from __future__ import annotations

import logging
import re
from datetime import UTC, datetime
from typing import Any

from googleapiclient.discovery import build  # type: ignore[import-untyped]

from agentgraph.auth.google_provider import (
    get_credentials as google_credentials,
)
from agentgraph.auth.google_provider import (
    get_user_email,
    list_google_accounts,
    verify_google_auth,
)
from agentgraph.connectors.base import (
    BaseConnector,
    ConnectorAccount,
    EdgeRecord,
    EntityBatch,
    EntityRecord,
    FetchPolicy,
    PersonRecord,
    ResourceType,
    SourceReference,
)
from agentgraph.graph.upsert import upsert_batch

logger = logging.getLogger(__name__)

_STALE_AFTER = 15 * 60
_GSHEETS_URL_RE = re.compile(
    r"https://docs\.google\.com/spreadsheets/d/(?P<spreadsheet_id>[a-zA-Z0-9_-]+)"
)


def _build_sheets_service(account_id: str | None = None) -> Any:
    return build("sheets", "v4", credentials=google_credentials(account_id))


def _build_drive_service(account_id: str | None = None) -> Any:
    return build("drive", "v3", credentials=google_credentials(account_id))


def _build_sheets_service_for(account_id: str | None) -> Any:
    return _build_sheets_service(account_id) if account_id is not None else _build_sheets_service()


def _build_drive_service_for(account_id: str | None) -> Any:
    return _build_drive_service(account_id) if account_id is not None else _build_drive_service()


def _web_url(spreadsheet_id: str) -> str:
    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/view"


def _download_url(spreadsheet_id: str) -> str:
    return (
        "https://www.googleapis.com/drive/v3/files/"
        f"{spreadsheet_id}/export?mimeType=application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _metadata(spreadsheet_id: str, account_id: str | None = None) -> dict[str, str | int | float | bool | None]:
    meta: dict[str, str | int | float | bool | None] = {
        "web_url": _web_url(spreadsheet_id),
        "download_url": _download_url(spreadsheet_id),
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    if account_id:
        meta["account_id"] = account_id
    return meta


def _extract_plain_text(spreadsheet: dict[str, Any], values_by_range: dict[str, list[list[str]]]) -> str:
    """Render each sheet as a labelled table of tab-separated rows."""
    sections: list[str] = []
    for sheet in spreadsheet.get("sheets", []):
        props = sheet.get("properties", {})
        sheet_title: str = props.get("title", "Sheet")
        rows = values_by_range.get(sheet_title, [])
        if not rows:
            continue
        lines = [sheet_title, "-" * len(sheet_title)]
        for row in rows:
            line = "\t".join(str(cell) for cell in row)
            if line.strip():
                lines.append(line)
        if len(lines) > 2:  # has content beyond header
            sections.append("\n".join(lines))
    return "\n\n".join(sections).strip()


class GoogleSheetsConnector(BaseConnector):
    source = "gsheets"
    fetch_policy = FetchPolicy(stale_after_seconds=_STALE_AFTER)
    url_patterns = ["https://docs.google.com/spreadsheets/*"]
    auth_label = "google"
    auth_description = "Google Sheets and uploaded .xlsx files: Spreadsheet entities with all tabs rendered as tab-separated cell text, plus owner authorship."

    @classmethod
    def run_auth_flow(cls, account_id: str | None = None, add: bool = False) -> None:
        from agentgraph_connector_google.auth import run_oauth_flow
        run_oauth_flow(account_id=account_id, add=add)

    @classmethod
    def get_authenticated_user(cls) -> str | None:
        return get_user_email()

    @classmethod
    def list_accounts(cls) -> list[ConnectorAccount]:
        return [
            ConnectorAccount(
                account_id=str(account["account_id"]),
                label=str(account["label"]),
                auth_group=cls.auth_label or cls.source,
                source=cls.source,
                user_id=account.get("email"),
                email=account.get("email"),
            )
            for account in list_google_accounts()
        ]

    @classmethod
    async def verify_auth(cls, account_id: str | None = None) -> tuple[str, str | None]:
        import asyncio
        return await asyncio.to_thread(verify_google_auth, account_id)

    @classmethod
    def current_user_ids(cls) -> list[str]:
        return [str(account["email"]) for account in list_google_accounts() if account.get("email")]

    def can_handle(self, url: str) -> bool:
        return self.resolve_url(url) is not None

    def resolve_url(self, url: str) -> SourceReference | None:
        match = _GSHEETS_URL_RE.match(url)
        if match is None:
            return None
        return SourceReference(
            source=self.source,
            resource_type="spreadsheet",
            resource_id=match.group("spreadsheet_id"),
        )

    def entity_url(self, platform_entity_id: str) -> str | None:
        return f"https://docs.google.com/spreadsheets/d/{platform_entity_id}"

    async def fetch(
        self,
        resource_type: ResourceType,
        resource_id: str,
        meta: dict[str, str] | None = None,
        account_id: str | None = None,
    ) -> EntityBatch:
        last_sync = await self.last_synced_at(resource_id)
        decision = self.fetch_policy.decide(last_sync)

        if decision == FetchPolicy.FRESH:
            logger.debug("gsheets/%s is fresh — updating last_accessed only", resource_id)
            await _touch_last_accessed(resource_id)
            return EntityBatch()

        logger.info("Fetching Google Sheet %s (policy=%s)", resource_id, decision)
        selected_account_id = account_id or ((meta or {}).get("account_id") if meta else None)
        batch = await _fetch_sheet(resource_id, account_id=selected_account_id)
        await upsert_batch(batch)
        return batch

    async def download(
        self,
        resource_type: ResourceType,
        resource_id: str,
        output_path: str | None = None,
    ) -> dict[str, Any]:
        from agentgraph_connector_google.gdrive import download_drive_file

        return await download_drive_file(resource_id, output_path)


async def _touch_last_accessed(spreadsheet_id: str) -> None:
    from agentgraph.core.context import get_backend

    await get_backend().touch_last_accessed("gsheets", spreadsheet_id)


def _extract_xlsx(data: bytes) -> tuple[str, str]:
    """Return (title, content) from raw xlsx bytes. title from first sheet name."""
    import io
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        return "", "(openpyxl not installed — cannot read Excel files)"

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sections: list[str] = []
    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        lines = [sheet.title, "-" * len(sheet.title)]
        for row in rows:
            line = "\t".join("" if cell is None else str(cell) for cell in row)
            if line.strip():
                lines.append(line)
        if len(lines) > 2:
            sections.append("\n".join(lines))
    title = wb.sheetnames[0] if wb.sheetnames else ""
    return title, "\n\n".join(sections).strip()


async def _fetch_excel_via_drive(
    spreadsheet_id: str,
    loop: Any,
    account_id: str | None = None,
) -> EntityBatch:
    """Download an Office-format file from Drive and parse as Excel."""
    import io

    from googleapiclient.http import MediaIoBaseDownload  # type: ignore[import-untyped]

    drive_service = await loop.run_in_executor(None, _build_drive_service_for, account_id)

    file_meta: dict[str, Any] = await loop.run_in_executor(
        None,
        lambda: drive_service.files().get(
            fileId=spreadsheet_id,
            fields="name,owners",
        ).execute(),
    )
    title: str = file_meta.get("name", "")

    def _download() -> bytes:
        buf = io.BytesIO()
        req = drive_service.files().get_media(fileId=spreadsheet_id)
        dl = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = dl.next_chunk()
        return buf.getvalue()

    raw = await loop.run_in_executor(None, _download)
    _, content = _extract_xlsx(raw)

    persons: list[PersonRecord] = []
    edges: list[EdgeRecord] = []
    for owner in file_meta.get("owners", []):
        email: str = owner.get("emailAddress", "")
        name: str = owner.get("displayName", "")
        if email:
            persons.append(PersonRecord(
                platform="gsheets",
                platform_user_id=email,
                canonical_email=email,
                display_name=name or None,
            ))
            edges.append(EdgeRecord(
                edge_type="authored",
                source_platform_user_id=email,
                target_platform_entity_id=spreadsheet_id,
                platform="gsheets",
            ))

    entity = EntityRecord(
        entity_type="Spreadsheet",
        platform="gsheets",
        platform_entity_id=spreadsheet_id,
        title=title,
        content=content,
        metadata=_metadata(spreadsheet_id, account_id),
        updated_at=datetime.now(UTC),
    )
    batch = EntityBatch(entities=[entity], persons=persons, edges=edges)
    batch.add_stubs_from(entity)
    return batch


async def _fetch_sheet(spreadsheet_id: str, account_id: str | None = None) -> EntityBatch:
    import asyncio

    from googleapiclient.errors import HttpError  # type: ignore[import-untyped]

    loop = asyncio.get_event_loop()
    sheets_service = await loop.run_in_executor(None, _build_sheets_service_for, account_id)
    file_meta: dict[str, Any] = {}

    try:
        spreadsheet: dict[str, Any] = await loop.run_in_executor(
            None,
            lambda: sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute(),
        )
    except HttpError as exc:
        if exc.status_code == 400 and "Office file" in str(exc):
            logger.info("gsheets/%s is an Office file — falling back to Drive download", spreadsheet_id)
            return await _fetch_excel_via_drive(spreadsheet_id, loop, account_id)
        raise

    title: str = spreadsheet.get("properties", {}).get("title", "")
    sheet_names: list[str] = [
        s["properties"]["title"]
        for s in spreadsheet.get("sheets", [])
        if s.get("properties", {}).get("title")
    ]

    # Fetch cell values for all sheets in one batch request
    values_by_range: dict[str, list[list[str]]] = {}
    if sheet_names:
        batch_result: dict[str, Any] = await loop.run_in_executor(
            None,
            lambda: sheets_service.spreadsheets().values().batchGet(
                spreadsheetId=spreadsheet_id,
                ranges=sheet_names,
            ).execute(),
        )
        for vr in batch_result.get("valueRanges", []):
            range_name: str = vr.get("range", "").split("!")[0].strip("'")
            values_by_range[range_name] = vr.get("values", [])

    content = _extract_plain_text(spreadsheet, values_by_range)

    persons: list[PersonRecord] = []
    edges: list[EdgeRecord] = []

    # Fetch owner via Drive API
    try:
        drive_service = await loop.run_in_executor(None, _build_drive_service_for, account_id)
        file_meta = await loop.run_in_executor(
            None,
            lambda: drive_service.files().get(
                fileId=spreadsheet_id,
                fields="owners,mimeType,webViewLink,webContentLink",
            ).execute(),
        )
        for owner in file_meta.get("owners", []):
            email: str = owner.get("emailAddress", "")
            name: str = owner.get("displayName", "")
            if email:
                persons.append(PersonRecord(
                    platform="gsheets",
                    platform_user_id=email,
                    canonical_email=email,
                    display_name=name or None,
                ))
                edges.append(EdgeRecord(
                    edge_type="authored",
                    source_platform_user_id=email,
                    target_platform_entity_id=spreadsheet_id,
                    platform="gsheets",
                ))
    except Exception:
        logger.debug("Could not fetch Drive ownership for sheet %s", spreadsheet_id)

    metadata: dict[str, str | int | float | bool | None] = {
        "web_url": file_meta.get("webViewLink") or _web_url(spreadsheet_id),
        "download_url": file_meta.get("webContentLink") or _download_url(spreadsheet_id),
        "mime_type": file_meta.get("mimeType") or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    if account_id:
        metadata["account_id"] = account_id

    entity = EntityRecord(
        entity_type="Spreadsheet",
        platform="gsheets",
        platform_entity_id=spreadsheet_id,
        title=title,
        content=content,
        metadata=metadata,
        updated_at=datetime.now(UTC),
    )

    batch = EntityBatch(entities=[entity], persons=persons, edges=edges)
    batch.add_stubs_from(entity)
    return batch
